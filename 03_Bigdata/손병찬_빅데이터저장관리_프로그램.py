# 손병찬_빅데이터저장관리_프로그램
import csv
import os
from openpyxl import load_workbook
import openpyxl

base_repository_name = 'Bigdata_Repository'
typeA_repository_name = 'TypeA'
typeB_repository_name = 'TypeB'
dir_delimeter = '/'
file_name = '시뮬레이션_대구광역시_관광지별_방문객'
file_format_A = 'xlsx'
file_format_B = 'csv'
simulation_count = 100
simulation_data = ['2341', '남구', '대구광역시', '충혼탑', '1', '1577']
typeA_size_limit = 10000
typeB_size_limit = 20000
is_header = False
is_first = False

def get_dest_file_name(file_index, repository_name, file_format, file_size_limit):
    global is_header
    dest_file_name = f'{base_repository_name}{dir_delimeter}{repository_name}{dir_delimeter}{file_name}{str(file_index)}.{file_format}'
    try:
        file_size = os.path.getsize(dest_file_name)
        if file_size > file_size_limit:
            dest_file_name = f'{base_repository_name}{dir_delimeter}{repository_name}{dir_delimeter}{file_name}{str(file_index+1)}.{file_format}'
            is_header = True
        else:
            is_header = False
    except:
        pass
    return dest_file_name

def save_csv(dest_file_name):
    global is_first, is_header
    csv_out_file = open(dest_file_name , 'a', newline='')
    filewriter = csv.writer(csv_out_file)
    if is_header == True or is_first == True:
        header_list = ['addrCd', 'gungu', 'sido', 'resNm', 'rnum', 'csForCnt']
        filewriter.writerow(header_list)
        is_first = False
        is_header = False
    for index in range(simulation_count):
        filewriter.writerow(simulation_data)
    csv_out_file.close()

def save_excel(dest_file_name):
    global is_first, is_header
    if is_header == True or is_first == True:
        header_list = ['addrCd', 'gungu', 'sido', 'resNm', 'rnum', 'csForCnt']
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = '수집_데이터'
        sheet.append(header_list)
        is_first = False
        is_header = False
    else:
        wb = load_workbook(filename=dest_file_name, read_only=False, data_only=False)
        sheet = wb.active
    for index in range(simulation_count):
        sheet.append(simulation_data)
    wb.save(dest_file_name)

def save_file(index, repository_name, file_format, file_size_limit):
    dest_file_name = get_dest_file_name(index, repository_name, file_format, file_size_limit)
    if file_format == 'csv':
        save_csv(dest_file_name)
    elif file_format == 'xlsx':
        save_excel(dest_file_name)
    print("\n===시뮬레이션 데이터가 정상 수집 및 저장되었습니다. 정상작동 완료===")

def file_count(repository_name):
    index = len(os.listdir((f'{base_repository_name}{dir_delimeter}{repository_name}')))
    return index

def make_dir(repository_name, file_format, file_size_limit):
    global is_first
    if not os.path.exists(base_repository_name):
        os.mkdir(base_repository_name)
    if not os.path.exists(f'{base_repository_name}{dir_delimeter}{repository_name}'):
        os.mkdir(f'{base_repository_name}{dir_delimeter}{repository_name}')
    if not os.path.exists(f'{base_repository_name}{dir_delimeter}{repository_name}{dir_delimeter}{file_name}1.{file_format}'):
        is_first = True
        save_file(1, repository_name, file_format, file_size_limit)
    else:
        save_file(file_count(repository_name), repository_name, file_format, file_size_limit)

def perform_work():
    while True:
        select_number = int(input("\n\t\t< 작업 수행 메뉴 >\
        \n\t\t1. Type A 데이터 수집\n\t\t2. Type B 데이터 수집\n\t\t3. 이전메뉴\
        \n\t\t메뉴를 선택하세요: "))
        if select_number == 1:
            make_dir(typeA_repository_name, file_format_A, typeA_size_limit )
        elif select_number == 2:
            make_dir(typeB_repository_name, file_format_B, typeB_size_limit)
        elif select_number == 3:
            return

def option():
    global base_repository_name, typeA_repository_name, file_format_A, typeA_size_limit
    global typeB_repository_name, typeB_size_limit, file_format_B, typeB_size_limit
    while True:
        print('''\n\t\t< 환경 설정 메뉴 >
        1. Base Repository: {}
        2. TypeA Repository명: {}
           TypeA 포멧: {}
        3. TypeA 데이터 용량제한: {}
        4. TypeB Repository명: {}
           TypeB 포멧: {}
        5. TypeB 데이터 용량제한: {}
        6. 이전메뉴'''.format(base_repository_name, typeA_repository_name, file_format_A
            , typeA_size_limit, typeB_repository_name, file_format_B, typeB_size_limit))
        select_number = int(input('\t\t메뉴를 선택하세요: '))
        if select_number == 1:
            base_repository_name = input('새로운 base repository 명을 입력하세요: ')
        elif select_number == 2:
            typeA_repository_name = input('Type A의 새로운 repository 명을 입력하세요: ')
            file_format_A = input('Type A의 새로운 포멧명을 입력하세요 ex) xlsx: ')
        elif select_number == 3:
            typeA_size_limit = int(input('Type A의 새로운 데이터용량 제한값을 입력하세요: '))
        elif select_number == 4:
            typeB_repository_name = input('Type B의 새로운 repository 명을 입력하세요: ')
            file_format_B = input('Type B의 새로운 포멧명을 입력하세요 ex) xlsx: ')
        elif select_number == 5:
            typeB_size_limit = int(input('Type B의 새로운 데이터용량 제한값을 입력하세요: '))
        elif select_number == 6:
            return

print("=경량화 빅데이터 저장소 시뮬레이션 v1.0 손병찬 =")

while True:
    select_main = int(input('''\n\t<메인메뉴>
    1. 환경설정
    2. 작업수행
    3. 종료
    메뉴를 선택하세요: '''))
    if select_main == 1:
        option()
    elif select_main == 2:
        perform_work()
    elif select_main == 3:
        break